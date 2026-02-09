from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from .models import Personel, AvansTalebi, AvansHareketi, IzinTalebi, MasrafTalebi, Zimmet, Egitim
from .forms import AvansTalepForm, PersonelForm, ExcelUploadForm, AvansIslemForm, PersonelKendiForm, IzinTalebiForm, IzinOnayForm, MasrafTalepForm, MasrafOnayForm, ZimmetForm, EgitimForm
from .models import yillik_izin_hakki_hesapla, calisma_gunleri_hesapla, get_profil_tipi
from django.contrib.auth.models import User
from django.contrib import messages
import csv
import openpyxl
from datetime import datetime


@login_required
def logout_view(request):
    """GET istegi ile cikis yap - CSRF gerektirmez"""
    from django.contrib.auth import logout
    logout(request)
    return redirect('login')


def is_staff(user):
    return user.is_staff


def is_muhasebe(user):
    """Kullanici muhasebe profili mi kontrol eder"""
    try:
        if user.is_staff:
            return True
        personel = Personel.objects.get(user=user)
        return personel.profil_tipi == 'Muhasebe'
    except Personel.DoesNotExist:
        return False


def is_patron(user):
    """Kullanici patron profili mi kontrol eder"""
    try:
        personel = Personel.objects.get(user=user)
        return personel.profil_tipi == 'Patron'
    except Personel.DoesNotExist:
        return False


def get_user_profile(request):
    """Kullanicinin profil tipini dondurur"""
    if not request.user.is_authenticated:
        return None
    try:
        personel = Personel.objects.get(user=request.user)
        return personel.profil_tipi
    except Personel.DoesNotExist:
        return None


def hareket_kaydet(personel, hareket_turu, miktar, aciklama=''):
    """Hareket kaydi olusturur"""
    onceki_borc = personel.guncel_avans_borcu
    if hareket_turu in ['TALEP', 'GIRIS']:
        yeni_borc = onceki_borc
    else:  # ODEME
        yeni_borc = onceki_borc - miktar
    
    AvansHareketi.objects.create(
        personel=personel,
        hareket_turu=hareket_turu,
        miktar=miktar,
        onceki_borc=onceki_borc,
        yeni_borc=yeni_borc,
        aciklama=aciklama
    )

@login_required
def index(request):
    """Ana sayfa yonlendirmesi - profil tipine gore"""
    user = request.user
    
    # Admin/Patron ise IK paneline yonlendir
    if user.is_staff:
        return redirect('ik_paneli')
    
    # Personel profilini kontrol et
    try:
        personel = Personel.objects.get(user=user)
        if personel.profil_tipi == 'Muhasebe':
            return redirect('ik_paneli')
        elif personel.profil_tipi == 'Patron':
            return redirect('ik_paneli')
        else:
            return redirect('personel_panel')
    except Personel.DoesNotExist:
        return redirect('personel_panel')

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u) or is_patron(u))
def ik_paneli(request):
    query = request.GET.get('q', '')
    
    if query:
        def turkish_lower(s):
            turkish_map = {
                'I': 'i', 'I': 'i', 'G': 'g', 'U': 'u', 'S': 's', 'O': 'o', 'C': 'c',
                'i': 'i', 'g': 'g', 'u': 'u', 's': 's', 'o': 'o', 'c': 'c'
            }
            result = ''
            for char in s:
                result += turkish_map.get(char, char.lower())
            return result
        
        query_lower = turkish_lower(query)
        
        all_personeller = Personel.objects.all()
        personeller = []
        for p in all_personeller:
            ad = turkish_lower(p.user.first_name)
            soyad = turkish_lower(p.user.last_name)
            tc = p.tc_no
            
            if (query_lower in ad or query_lower in soyad or query in tc):
                personeller.append(p)
    else:
        personeller = Personel.objects.all()
    
    return render(request, 'core/ik_paneli.html', {'personeller': personeller})

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def muhasebe_paneli(request):
    bekleyen_talepler = AvansTalebi.objects.filter(durum='Bekliyor')
    onaylanan_talepler = AvansTalebi.objects.filter(durum='Onaylandi').order_by('-tarih')[:10]
    reddedilen_talepler = AvansTalebi.objects.filter(durum='Reddedildi').order_by('-tarih')[:10]
    personeller = Personel.objects.all()
    return render(request, 'core/muhasebe_paneli.html', {
        'bekleyen_talepler': bekleyen_talepler,
        'onaylanan_talepler': onaylanan_talepler,
        'reddedilen_talepler': reddedilen_talepler,
        'personeller': personeller
    })

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def personel_detay(request, personel_id):
    personel = get_object_or_404(Personel, id=personel_id)
    talepler = AvansTalebi.objects.filter(personel=personel).order_by('-tarih')
    hareketler = AvansHareketi.objects.filter(personel=personel).order_by('-tarih')
    
    if request.method == 'POST':
        form = AvansIslemForm(request.POST)
        if form.is_valid():
            islem_turu = form.cleaned_data['islem_turu']
            miktar = form.cleaned_data['miktar']
            aciklama = form.cleaned_data['aciklama']
            
            onceki_borc = personel.guncel_avans_borcu
            
            if islem_turu == 'EKLE':
                personel.guncel_avans_borcu += miktar
                hareket_turu = 'ODEME'
            else:  # AZALT
                personel.guncel_avans_borcu -= miktar
                hareket_turu = 'GIRIS'
            
            personel.save()
            
            AvansHareketi.objects.create(
                personel=personel,
                hareket_turu=hareket_turu,
                miktar=miktar,
                onceki_borc=onceki_borc,
                yeni_borc=personel.guncel_avans_borcu,
                aciklama=aciklama or f"Borc eklendi" if islem_turu == 'EKLE' else "Borc azaltildi"
            )
            
            messages.success(request, f'Avans islemi basariyla uygulandi. Yeni borc: {personel.guncel_avans_borcu} TL')
            return redirect('personel_detay', personel_id=personel_id)
    else:
        form = AvansIslemForm()
    
    return render(request, 'core/personel_detay.html', {
        'personel': personel,
        'talepler': talepler,
        'hareketler': hareketler,
        'islem_form': form
    })

@login_required
def personel_panel(request):
    personel = get_object_or_404(Personel, user=request.user)
    talepler = AvansTalebi.objects.filter(personel=personel).order_by('-tarih')
    izin_talepleri = IzinTalebi.objects.filter(personel=personel).order_by('-talep_tarihi')
    
    bilgi_form = None
    if request.method == 'POST' and 'bilgi_guncelle' in request.POST:
        bilgi_form = PersonelKendiForm(request.POST, instance=personel)
        if bilgi_form.is_valid():
            bilgi_form.save()
            messages.success(request, 'Bilgileriniz guncellendi.')
            return redirect('personel_panel')
    else:
        bilgi_form = PersonelKendiForm(instance=personel)
    
    if request.method == 'POST' and 'avans_talep' in request.POST:
        form = AvansTalepForm(request.POST)
        if form.is_valid():
            talep = form.save(commit=False)
            talep.personel = personel
            hareket_kaydet(personel, 'TALEP', talep.miktar, talep.aciklama)
            talep.save()
            messages.success(request, 'Avans talebiniz iletildi.')
            return redirect('personel_panel')
    else:
        form = AvansTalepForm()
        
    return render(request, 'core/personel.html', {
        'personel': personel,
        'talepler': talepler,
        'izin_talepleri': izin_talepleri,
        'form': form,
        'bilgi_form': bilgi_form
    })

@login_required
@user_passes_test(is_staff)
def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="avans_borclari.csv"'
    response.write(u'\ufeff'.encode('utf8'))
    
    writer = csv.writer(response)
    writer.writerow(['Ad', 'Soyad', 'TC No', 'Borc Miktari'])
    
    borclular = Personel.objects.exclude(guncel_avans_borcu=0)
    for p in borclular:
        writer.writerow([p.user.first_name, p.user.last_name, p.tc_no, p.guncel_avans_borcu])
        
    return response

@login_required
@user_passes_test(is_staff)
def borc_guncelle(request, personel_id):
    personel = get_object_or_404(Personel, id=personel_id)
    if request.method == 'POST':
        yeni_borc = request.POST.get('guncel_avans_borcu')
        personel.guncel_avans_borcu = yeni_borc
        personel.save()
        messages.success(request, f'{personel.user.get_full_name()} borcu guncellendi.')
    return redirect('muhasebe_paneli')

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def avans_islem(request, talep_id, islem):
    talep = get_object_or_404(AvansTalebi, id=talep_id)
    if islem == 'onayla':
        talep.durum = 'Onaylandi'
        onceki_borc = talep.personel.guncel_avans_borcu
        talep.personel.guncel_avans_borcu += talep.miktar
        talep.personel.save()
        
        AvansHareketi.objects.create(
            personel=talep.personel,
            hareket_turu='ODEME',
            miktar=talep.miktar,
            onceki_borc=onceki_borc,
            yeni_borc=talep.personel.guncel_avans_borcu,
            aciklama=f'Avans talebi onaylandi: {talep.miktar} TL'
        )
        messages.success(request, 'Talep onaylandi ve borc kaydedildi.')
    elif islem == 'reddet':
        talep.durum = 'Reddedildi'
        messages.warning(request, 'Talep reddedildi.')
    talep.save()
    return redirect('muhasebe_paneli')

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def personel_ekle(request):
    if request.method == 'POST':
        form = PersonelForm(request.POST)
        if form.is_valid():
            tc_no = form.cleaned_data['tc_no']
            
            if Personel.objects.filter(tc_no=tc_no).exists():
                messages.error(request, 'Bu TC Kimlik Numarasi zaten kayitli!')
                return render(request, 'core/personel_ekle.html', {'form': form})
            
            first_name = form.cleaned_data.get('first_name', '')
            last_name = form.cleaned_data.get('last_name', '')
            
            # Türkçe karakterleri İngilizce karakterlere çevir
            def turkish_to_english(text):
                turkish_map = {
                    'ğ': 'g', 'Ğ': 'G',
                    'ş': 's', 'Ş': 'S',
                    'ı': 'i', 'İ': 'I',
                    'ö': 'o', 'Ö': 'O',
                    'ü': 'u', 'Ü': 'U',
                    'ç': 'c', 'Ç': 'C'
                }
                return ''.join(turkish_map.get(c, c) for c in text)
            
            first_name_en = turkish_to_english(first_name.lower())
            last_name_en = turkish_to_english(last_name.lower())
            email = f"{first_name_en}.{last_name_en}@ets-telekom.com.tr"
            
            username = tc_no
            password = tc_no[-5:] if len(tc_no) >= 5 else tc_no
            
            user = User.objects.create_user(
                username=username, 
                password=password, 
                email=email, 
                first_name=first_name, 
                last_name=last_name
            )
            
            personel = form.save(commit=False)
            personel.user = user
            
            if not personel.kalan_izin:
                personel.kalan_izin = 0
            if personel.telefon is None or personel.telefon == '':
                personel.telefon = None
            
            personel.izin_hakki = yillik_izin_hakki_hesapla(personel.ise_giris_tarihi, personel.dogum_tarihi)
            
            personel.save()
            
            messages.success(request, f'Personel basariyla eklendi!\nKullanici Adi: {username}\nSifre: {password}')
            return redirect('ik_paneli')
    else:
        form = PersonelForm()
    return render(request, 'core/personel_ekle.html', {'form': form})

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def personel_duzenle(request, personel_id):
    personel = get_object_or_404(Personel, id=personel_id)
    
    def turkish_to_english(text):
        """Türkçe karakterleri İngilizce karakterlere çevirir"""
        turkish_map = {
            'ğ': 'g', 'Ğ': 'G',
            'ş': 's', 'Ş': 'S',
            'ı': 'i', 'İ': 'I',
            'ö': 'o', 'Ö': 'O',
            'ü': 'u', 'Ü': 'U',
            'ç': 'c', 'Ç': 'C'
        }
        return ''.join(turkish_map.get(c, c) for c in text)
    
    if request.method == 'POST':
        form = PersonelForm(request.POST, instance=personel)
        if form.is_valid():
            first_name = form.cleaned_data.get('first_name', '')
            last_name = form.cleaned_data.get('last_name', '')
            
            # Email otomatik oluştur (Türkçe karakterleri İngilizce yap)
            first_name_en = turkish_to_english(first_name.lower())
            last_name_en = turkish_to_english(last_name.lower())
            email = f"{first_name_en}.{last_name_en}@ets-telekom.com.tr"
            
            personel.user.first_name = first_name
            personel.user.last_name = last_name
            personel.user.email = email
            personel.user.save()
            
            personel = form.save(commit=False)
            if not personel.kalan_izin:
                personel.kalan_izin = 0
            if personel.telefon is None or personel.telefon == '':
                personel.telefon = None
            
            personel.izin_hakki = yillik_izin_hakki_hesapla(personel.ise_giris_tarihi, personel.dogum_tarihi)
            
            personel.save()
            messages.success(request, 'Personel bilgileri guncellendi.')
            return redirect('ik_paneli')
    else:
        initial_data = {
            'first_name': personel.user.first_name,
            'last_name': personel.user.last_name,
            'email': personel.user.email,
        }
        if personel.ise_giris_tarihi:
            initial_data['ise_giris_tarihi'] = personel.ise_giris_tarihi.strftime('%Y-%m-%d')
        form = PersonelForm(instance=personel, initial=initial_data)
    return render(request, 'core/personel_duzenle.html', {'form': form, 'personel': personel})

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def personel_sil(request, personel_id):
    personel = get_object_or_404(Personel, id=personel_id)
    user = personel.user
    personel.delete()
    user.delete()
    messages.error(request, 'Personel silindi.')
    return redirect('ik_paneli')

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def personel_yukle(request):
    """Excel ile toplu personel ekleme - Sadeleştirilmiş"""
    def turkish_to_english(text):
        """Türkçe karakterleri İngilizce karakterlere çevirir"""
        if not text:
            return ''
        turkish_map = {
            'ğ': 'g', 'Ğ': 'G',
            'ş': 's', 'Ş': 'S',
            'ı': 'i', 'İ': 'I',
            'ö': 'o', 'Ö': 'O',
            'ü': 'u', 'Ü': 'U',
            'ç': 'c', 'Ç': 'C'
        }
        return ''.join(turkish_map.get(c, c) for c in str(text))
    
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            
            sheet = wb.active
            count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                
                try:
                    # Excel format: Ad, Soyad, TC No, Telefon, Ise Giris, Dogum Tarihi
                    first_name = row[0] if row[0] else ''
                    last_name = row[1] if row[1] else ''
                    tc_no = str(row[2]) if row[2] else ''
                    telefon = row[3] if row[3] else None
                    giris_tarihi = row[4] if row[4] else None
                    dogum_tarihi = row[5] if row[5] else None
                    
                    if not tc_no:
                        continue
                    
                    if Personel.objects.filter(tc_no=tc_no).exists():
                        continue
                    
                    # Email otomatik oluştur
                    first_name_en = turkish_to_english(first_name.lower())
                    last_name_en = turkish_to_english(last_name.lower())
                    email = f"{first_name_en}.{last_name_en}@ets-telekom.com.tr"
                    
                    username = tc_no
                    password = tc_no[-5:] if len(tc_no) >= 5 else tc_no
                    
                    user = User.objects.create_user(
                        username=username, 
                        password=password, 
                        email=email, 
                        first_name=first_name, 
                        last_name=last_name
                    )
                    
                    # Tarih parse et
                    def parse_date(date_val, default_days_ago=365):
                        if not date_val:
                            from datetime import timedelta
                            return datetime.today().date() - timedelta(days=default_days_ago)
                        if isinstance(date_val, str):
                            try:
                                return datetime.strptime(date_val, '%d.%m.%Y').date()
                            except ValueError:
                                try:
                                    return datetime.strptime(date_val, '%Y-%m-%d').date()
                                except ValueError:
                                    from datetime import timedelta
                                    return datetime.today().date() - timedelta(days=default_days_ago)
                        elif isinstance(date_val, datetime):
                            return date_val.date()
                        return date_val
                    
                    giris_date = parse_date(giris_tarihi)
                    dogum_date = parse_date(dogum_tarihi, default_days_ago=365*20)  # 20 yıl önce
                    
                    # İzin hakkını otomatik hesapla
                    izin_hakki = yillik_izin_hakki_hesapla(giris_date, dogum_date)
                    
                    Personel.objects.create(
                        user=user,
                        tc_no=tc_no,
                        telefon=telefon,
                        ise_giris_tarihi=giris_date,
                        dogum_tarihi=dogum_date,
                        izin_hakki=izin_hakki,
                        kalan_izin=izin_hakki,
                        guncel_avans_borcu=0
                    )
                    count += 1
                except Exception as e:
                    messages.error(request, f"Hata ({row[0]} {row[1]}): {str(e)}")
                    
            messages.success(request, f'{count} personel basariyla yuklendi.')
            return redirect('ik_paneli')
    else:
        form = ExcelUploadForm()
    return render(request, 'core/personel_yukle.html', {'form': form})

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def download_excel_template(request):
    """Sadeleştirilmiş Excel şablonu indir"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personel Sablon"
    
    # Yeni sadeleştirilmiş format
    headers = ['Ad', 'Soyad', 'TC No', 'Telefon', 'Ise Giris (GG.AA.YYYY)', 'Dogum Tarihi (GG.AA.YYYY)']
    ws.append(headers)
    
    ws.append(['Ahmet', 'Yilmaz', '12345678901', '05551234567', '01.08.2020', '01.01.1990'])
    ws.append(['Ayse', 'Demir', '10987654321', '05559876543', '15.03.2021', '20.06.1995'])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="personel_yukleme_sablonu.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def toplu_excel_indir(request):
    """Tüm personeli Excel formatında indir"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personel Listesi"
    
    # Başlık satırı
    headers = ['Ad', 'Soyad', 'TC No', 'Telefon', 'E-posta', 'Ise Giris', 'Dogum Tarihi', 'Izin Hakki', 'Kalan Izin', 'Avans Borcu', 'Profil Tipi']
    ws.append(headers)
    
    # Tüm personeli getir
    personeller = Personel.objects.select_related('user').all().order_by('user__first_name')
    
    for p in personeller:
        ws.append([
            p.user.first_name,
            p.user.last_name,
            p.tc_no,
            p.telefon,
            p.user.email,
            p.ise_giris_tarihi.strftime('%d.%m.%Y') if p.ise_giris_tarihi else '',
            p.dogum_tarihi.strftime('%d.%m.%Y') if p.dogum_tarihi else '',
            p.izin_hakki,
            p.kalan_izin,
            p.guncel_avans_borcu,
            p.profil_tipi
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="personel_listesi.xlsx"'
    wb.save(response)
    return response

@login_required
def sifre_degistir(request):
    if request.method == 'POST':
        yeni_sifre = request.POST.get('yeni_sifre')
        if yeni_sifre:
            request.user.set_password(yeni_sifre)
            request.user.save()
            messages.success(request, 'Sifreniz basariyla degistirildi.')
            return redirect('personel_panel')
        else:
            messages.error(request, 'Sifre bos olamaz!')
    return render(request, 'core/sifre_degistir.html')

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def sifre_resetle(request, personel_id):
    personel = get_object_or_404(Personel, id=personel_id)
    yeni_sifre = personel.tc_no[-5:] if len(personel.tc_no) >= 5 else personel.tc_no
    personel.user.set_password(yeni_sifre)
    personel.user.save()
    messages.success(request, f'{personel.user.get_full_name()} sifresi TC No son 5 hanesine resetlendi.')
    return redirect('personel_duzenle', personel_id)


# ========== IZIN TALEPLERI ==========

@login_required
def izin_talep_et(request):
    """Personelin izin talep etmesi"""
    personel = get_object_or_404(Personel, user=request.user)
    
    if request.method == 'POST':
        form = IzinTalebiForm(request.POST)
        if form.is_valid():
            talep = form.save(commit=False)
            talep.personel = personel
            
            # Tarih araligindan gun sayisini hesapla
            talep.gun_sayisi = calisma_gunleri_hesapla(talep.baslangic_tarihi, talep.bitis_tarihi)
            
            talep.save()
            messages.success(request, 'Izin talebiniz basariyla olusturuldu.')
            return redirect('personel_panel')
    else:
        form = IzinTalebiForm()
    
    return render(request, 'core/izin_talep.html', {'form': form})

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def izin_listesi(request):
    """Admin veya Muhasebe icin tum talepleri goster"""
    # Bekleyen izin talepleri
    bekleyen_izin_talepleri = IzinTalebi.objects.filter(durum='Bekliyor').order_by('-talep_tarihi')
    # Bekleyen avans talepleri
    bekleyen_avans_talepleri = AvansTalebi.objects.filter(durum='Bekliyor').order_by('-tarih')
    
    return render(request, 'core/izin_listesi.html', {
        'bekleyen_izin_talepleri': bekleyen_izin_talepleri,
        'bekleyen_avans_talepleri': bekleyen_avans_talepleri
    })

@login_required
@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u))
def izin_islem(request, talep_id, islem):
    """Izin talebini onayla veya reddet"""
    talep = get_object_or_404(IzinTalebi, id=talep_id)
    
    if islem == 'onayla':
        talep.durum = 'Onaylandi'
        messages.success(request, 'Izin talebi onaylandi.')
    elif islem == 'reddet':
        talep.durum = 'Reddedildi'
        messages.warning(request, 'Izin talebi reddedildi.')
    elif islem == 'iptal':
        talep.durum = 'Iptal Edildi'
        messages.info(request, 'Izin talebi iptal edildi.')
    
    talep.save()
    return redirect('izin_listesi')

@login_required
def izin_detay(request, talep_id):
    """Izin talebi detayini goster"""
    talep = get_object_or_404(IzinTalebi, id=talep_id)
    return render(request, 'core/izin_detay.html', {'talep': talep})

def login_view(request):
    """Kullanici girisi"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            
            # Profil tipine gore yonlendir
            if user.is_staff:
                return redirect('ik_paneli')
            
            try:
                personel = Personel.objects.get(user=user)
                if personel.profil_tipi == 'Muhasebe':
                    return redirect('muhasebe_paneli')
                elif personel.profil_tipi == 'Patron':
                    return redirect('ik_paneli')
            except Personel.DoesNotExist:
                pass
            
            return redirect('index')
        else:
            messages.error(request, 'Kullanici adi veya sifre yanlis!')
    
    return render(request, 'core/login.html')

def register_view(request):
    """Yeni kullanici kaydi"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Bu kullanici adi zaten alinmis!')
        else:
            user = User.objects.create_user(
                username=username,
                password=password,
                email=email,
                first_name=first_name,
                last_name=last_name
            )
            
            Personel.objects.create(
                user=user,
                tc_no=username,
                profil_tipi='Personel'
            )
            
            login(request, user)
            messages.success(request, 'Kayit basarili! Hosgeldiniz.')
            return redirect('index')
    
    return render(request, 'core/register.html')


@login_required
def hesapla_gun(request):
    """Calisma gunu hesaplama (placeholder)"""
    if request.method == 'POST':
        baslangic = request.POST.get('baslangic')
        bitis = request.POST.get('bitis')
        if baslangic and bitis:
            try:
                from datetime import datetime
                b_date = datetime.strptime(baslangic, '%Y-%m-%d').date()
                e_date = datetime.strptime(bitis, '%Y-%m-%d').date()
                gun = calisma_gunleri_hesapla(b_date, e_date)
                messages.info(request, f'Bu tarih araliginda {gun} is gunu var.')
            except Exception as e:
                messages.error(request, f'Hata: {str(e)}')
        else:
            messages.warning(request, 'Lutfen baslangic ve bitis tarihi secin.')
    return redirect('izin_talep_et')


@login_required
def zimmetlerim(request):
    """Personelin zimmetlerini goster"""
    personel = get_object_or_404(Personel, user=request.user)
    return render(request, 'core/zimmetlerim.html', {'personel': personel})


@login_required
def egitimlerim(request):
    """Personelin egitimlerini goster"""
    personel = get_object_or_404(Personel, user=request.user)
    return render(request, 'core/egitimlerim.html', {'personel': personel})


# ==================== MASRAF TALEBI VIEWS ====================

@login_required
def masraf_talep_et(request):
    """Personelin masraf talep etmesi"""
    personel = get_object_or_404(Personel, user=request.user)
    
    if request.method == 'POST':
        form = MasrafTalepForm(request.POST, request.FILES)
        if form.is_valid():
            masraf = form.save(commit=False)
            masraf.personel = personel
            masraf.save()
            messages.success(request, 'Masraf talebiniz başarıyla oluşturuldu.')
            return redirect('masraf_listem')
    else:
        form = MasrafTalepForm()
    
    return render(request, 'core/masraf_talep.html', {'form': form, 'personel': personel})


@login_required
def masraf_listem(request):
    """Personelin kendi masraf taleplerini görmesi"""
    personel = get_object_or_404(Personel, user=request.user)
    masraflar = MasrafTalebi.objects.filter(personel=personel).order_by('-talep_tarihi')
    
    return render(request, 'core/masraf_listem.html', {'masraflar': masraflar, 'personel': personel})


@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u) or is_patron(u))
def masraf_listesi(request):
    """Admin/Muhasebe/Patron için tüm masraf talepleri"""
    durum_filter = request.GET.get('durum', '')
    
    masraflar = MasrafTalebi.objects.all().order_by('-talep_tarihi')
    
    if durum_filter:
        masraflar = masraflar.filter(durum=durum_filter)
    
    return render(request, 'core/masraf_yonetim.html', {'masraflar': masraflar, 'durum_filter': durum_filter})


@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u) or is_patron(u))
def masraf_islem(request, talep_id, islem):
    """Masraf talebini onayla veya reddet"""
    talep = get_object_or_404(MasrafTalebi, id=talep_id)
    
    if islem == 'onayla':
        talep.durum = 'Onaylandı'
        talep.admin_notu = request.GET.get('not', '')
        talep.save()
        messages.success(request, f'{talep.personel.user.get_full_name()} için masraf talebi onaylandı.')
    elif islem == 'reddet':
        talep.durum = 'Reddedildi'
        talep.admin_notu = request.GET.get('not', '')
        talep.save()
        messages.warning(request, f'{talep.personel.user.get_full_name()} için masraf talebi reddedildi.')
    
    return redirect('masraf_yonetim')


@login_required
def masraf_detay(request, talep_id):
    """Masraf talebi detayı"""
    talep = get_object_or_404(MasrafTalebi, id=talep_id)
    
    # Admin/muhasebe/patron doğrudan görebilir
    if is_staff(request.user) or is_muhasebe(request.user) or is_patron(request.user):
        return render(request, 'core/masraf_detay.html', {'talep': talep})
    
    # Personel sadece kendi talebini görebilir
    personel = get_object_or_404(Personel, user=request.user)
    if talep.personel != personel:
        messages.error(request, 'Bu talebi görme yetkiniz yok.')
        return redirect('personel_panel')
    
    return render(request, 'core/masraf_detay.html', {'talep': talep, 'personel': personel})


@login_required
def avans_taleplerim(request):
    """Personelin avans taleplerini gördüğü sayfa"""
    personel = get_object_or_404(Personel, user=request.user)
    avans_talepleri = AvansTalebi.objects.filter(personel=personel).order_by('-tarih')
    
    form = None
    if request.method == 'POST' and 'avans_talep' in request.POST:
        form = AvansTalepForm(request.POST)
        if form.is_valid():
            talep = form.save(commit=False)
            talep.personel = personel
            hareket_kaydet(personel, 'TALEP', talep.miktar, talep.aciklama)
            talep.save()
            messages.success(request, 'Avans talebiniz iletildi.')
            return redirect('avans_taleplerim')
    else:
        form = AvansTalepForm()
    
    return render(request, 'core/avans_taleplerim.html', {'avans_talepleri': avans_talepleri, 'form': form})

@login_required
def izin_taleplerim(request):
    """Personelin izin taleplerini gördüğü sayfa"""
    personel = get_object_or_404(Personel, user=request.user)
    izin_talepleri = IzinTalebi.objects.filter(personel=personel).order_by('-talep_tarihi')
    return render(request, 'core/izin_taleplerim.html', {'izin_talepleri': izin_talepleri, 'personel': personel})

@login_required
def masraf_taleplerim(request):
    """Personelin masraf taleplerini gördüğü sayfa"""
    personel = get_object_or_404(Personel, user=request.user)
    masraf_talep = MasrafTalebi.objects.filter(personel=personel).order_by('-talep_tarihi')
    return render(request, 'core/masraf_taleplerim.html', {'masraf_talep': masraf_talep})

@login_required
def taleplerim(request):
    """Personelin tüm taleplerini gördüğü sayfa"""
    personel = get_object_or_404(Personel, user=request.user)
    
    # Tüm talepleri getir
    avans_talepleri = AvansTalebi.objects.filter(personel=personel).order_by('-tarih')
    izin_talepleri = IzinTalebi.objects.filter(personel=personel).order_by('-talep_tarihi')
    masraf_talepleri = MasrafTalebi.objects.filter(personel=personel).order_by('-talep_tarihi')
    
    context = {
        'personel': personel,
        'avans_talepleri': avans_talepleri,
        'izin_talepleri': izin_talepleri,
        'masraf_talepleri': masraf_talepleri,
    }
    
    return render(request, 'core/taleplerim.html', context)


@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u) or is_patron(u))
def talep_yonetim(request):
    """Admin/Muhasebe/Patron için tüm taleplerin yönetim paneli"""
    # Filtreler
    tur_filter = request.GET.get('tur', '')
    durum_filter = request.GET.get('durum', '')
    
    # Tüm talepleri getir
    avans_talepleri = AvansTalebi.objects.all().order_by('-tarih')
    izin_talepleri = IzinTalebi.objects.all().order_by('-talep_tarihi')
    masraf_talepleri = MasrafTalebi.objects.all().order_by('-talep_tarihi')
    
    # Durum filtresi
    if durum_filter:
        avans_talepleri = avans_talepleri.filter(durum=durum_filter)
        izin_talepleri = izin_talepleri.filter(durum=durum_filter)
        masraf_talepleri = masraf_talepleri.filter(durum=durum_filter)
    
    context = {
        'avans_talepleri': avans_talepleri,
        'izin_talepleri': izin_talepleri,
        'masraf_talepleri': masraf_talepleri,
        'tur_filter': tur_filter,
        'durum_filter': durum_filter,
    }
    
    return render(request, 'core/talep_yonetim.html', context)


# ========== ZIMMET VE EGITIM YONETIMI ==========

@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u) or is_patron(u))
def zimmet_yonetim_home(request):
    """Zimmet yönetimi için personel seçimi"""
    query = request.GET.get('q', '')
    if query:
        personeller = Personel.objects.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(tc_no__icontains=query)
        ).order_by('user__first_name')
    else:
        personeller = Personel.objects.all().order_by('user__first_name')[:50]
    return render(request, 'core/zimmet_yonetim_home.html', {'personeller': personeller, 'query': query})


@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u) or is_patron(u))
def egitim_yonetim_home(request):
    """Eğitim yönetimi için personel seçimi"""
    query = request.GET.get('q', '')
    if query:
        personeller = Personel.objects.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(tc_no__icontains=query)
        ).order_by('user__first_name')
    else:
        personeller = Personel.objects.all().order_by('user__first_name')[:50]
    return render(request, 'core/egitim_yonetim_home.html', {'personeller': personeller, 'query': query})


@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u) or is_patron(u))
def zimmet_yonetimi(request, personel_id):
    """Personel zimmetlerini yönetme sayfası"""
    personel = get_object_or_404(Personel, id=personel_id)
    zimmeter = Zimmet.objects.filter(personel=personel).order_by('-tarih')
    
    if request.method == 'POST':
        form = ZimmetForm(request.POST)
        if form.is_valid():
            zimmet = form.save(commit=False)
            zimmet.personel = personel
            zimmet.save()
            messages.success(request, 'Zimmet başarıyla eklendi.')
            return redirect('zimmet_yonetimi', personel_id)
    else:
        form = ZimmetForm()
    
    return render(request, 'core/zimmet_yonetimi.html', {
        'personel': personel,
        'zimmeter': zimmeter,
        'form': form
    })


@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u) or is_patron(u))
def zimmet_sil(request, zimmet_id):
    """Zimmet silme"""
    zimmet = get_object_or_404(Zimmet, id=zimmet_id)
    personel_id = zimmet.personel.id
    zimmet.delete()
    messages.success(request, 'Zimmet başarıyla silindi.')
    return redirect('zimmet_yonetimi', personel_id)


@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u) or is_patron(u))
def egitim_yonetimi(request, personel_id):
    """Personel eğitimlerini yönetme sayfası"""
    personel = get_object_or_404(Personel, id=personel_id)
    egitimler = Egitim.objects.filter(personel=personel).order_by('-tarih')
    
    if request.method == 'POST':
        form = EgitimForm(request.POST)
        if form.is_valid():
            egitim = form.save(commit=False)
            egitim.personel = personel
            egitim.save()
            messages.success(request, 'Eğitim başarıyla eklendi.')
            return redirect('egitim_yonetimi', personel_id)
    else:
        form = EgitimForm()
    
    return render(request, 'core/egitim_yonetimi.html', {
        'personel': personel,
        'egitimler': egitimler,
        'form': form
    })


@user_passes_test(lambda u: is_staff(u) or is_muhasebe(u) or is_patron(u))
def egitim_sil(request, egitim_id):
    """Eğitim silme"""
    egitim = get_object_or_404(Egitim, id=egitim_id)
    personel_id = egitim.personel.id
    egitim.delete()
    messages.success(request, 'Eğitim başarıyla silindi.')
    return redirect('egitim_yonetimi', personel_id)
